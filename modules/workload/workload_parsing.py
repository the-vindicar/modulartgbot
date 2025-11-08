"""Реализует работу с Excel файлами, содержащими нагрузку преподавателей."""
import typing as t
from collections import defaultdict
import dataclasses
from enum import StrEnum

from openpyxl.worksheet.worksheet import Worksheet


__all__ = [
    'WorkloadType', 'Activity', 'EducationType', 'ExamType', 'LoadUnit',
    'TeacherWorkload', 'parse_workload',
]


class WorkloadType(StrEnum):
    """Типы занятости преподавателя (штат, почасовка, совмещение и т.п.)"""
    MAIN = 'ш'
    HOURLY = 'h'
    INTERNAL = 'i'
    G = 'g'
    _order_ = ('MAIN', 'HOURLY', 'INTERNAL', 'G')


class Activity(StrEnum):
    """Виды деятельности"""
    LECTURE = 'Лек'
    PRACTICAL = 'Пр'
    LAB = 'Лаб'
    YEARLYTASK = 'КР'
    YEARLYPROJECT = 'КП'
    INTERNSHIP = 'Прак'
    GRADPROJECT = 'ВКР'
    MAGGRADPROJECT = 'ДРМ'
    MAGISTER = 'РукМаг'
    GRADCOMITTEE = 'ГАК'
    _order_ = ('LECTURE', 'PRACTICAL', 'LAB', 'YEARLYTASK', 'YEARLYPROJECT', 'INTERNSHIP',
               'GRADPROJECT', 'MAGGRADPROJECT', 'MAGISTER', 'GRADCOMITTEE')


class EducationType(StrEnum):
    """Тип обучения: очное, заочное и т.п."""
    DAYHOURS = 'Очная форма'
    OFFHOURS = 'Заочная форма'
    COMBINED = 'Очно-заочная форма'


class ExamType(StrEnum):
    """Тип контроля: зачёт, экзамен и т.п."""
    PASS = 'Зч'
    DIFFPASS = 'ЗчО'
    EXAM = 'Эк'


@dataclasses.dataclass
class LoadUnit:
    """Описание одного вида деятельности преподавателя."""
    course: str
    group: str
    subgroup: t.Optional[int]
    education_type: t.Optional[EducationType]
    education_level: str
    activity: t.Union[Activity, str]
    exam: t.Optional[ExamType]
    year: int
    semester: t.Optional[int]
    student_count: int
    week_count: t.Optional[int]
    worktime_hours: float
    other_hours: float


TeacherWorkload = dict[WorkloadType, list[LoadUnit]]


def find_column(ws: Worksheet, *targets: tuple[str, int]) -> int:
    """Находит столбец, содержащий в заданной строке заданное значение, и возвращает его номер.
    Выбрасывает IndexError, если такого значения нет.

    :param ws: Лист, на котором ищем значения.
    :param targets: Последовательность пар "значение-строка". Засчитывается соответствие любой паре."""
    for c in range(1, ws.max_column + 1):
        for value, row in targets:
            cell = ws.cell(row=row, column=c)
            if cell.value and str(cell.value) == value:
                return c
    raise IndexError()


def parse_workload(ws: Worksheet) -> dict[str, TeacherWorkload]:
    """Читает таблицу с нагрузкой и разделяет нагрузку по преподавателям."""
    lesson_col = find_column(ws, ('6', 5))
    year_col = find_column(ws, ('8', 5))
    group_col = find_column(ws, ('Группа', 1), ('9', 5))
    students_col = find_column(ws, ('Кол-во студентов', 1), ('10', 5))
    weeks_col = find_column(ws, ('Недель', 1), ('11', 5))
    activity_col = find_column(ws, ('Вид занятий', 1), ('12', 5))
    exam_col = find_column(ws, ('Виды контроля', 1), ('14', 5))
    worktime_col = find_column(ws, ('23', 5))
    othertime_col = find_column(ws, ('24', 5))
    teacher_col = find_column(ws, ('Преподаватель', 1), ('27', 5))
    edu_level_col = find_column(ws, ('Уровень', 1), ('45', 5))
    teachtype_col = find_column(ws, ('Форма обучения', 1), ('46', 5))

    teacher_workload = defaultdict(lambda: defaultdict(list))
    header_max_row = max(cellrange.max_row for cellrange in ws.merged_cells.ranges if cellrange.min_row == 1)
    row = header_max_row + 2
    while True:
        if not ws.cell(row=row, column=1).value:
            break
        cell = ws.cell(row=row, column=teacher_col)
        if cell.data_type != 's' or not cell.value:
            row += 1
            continue
        teacher_name = cell.value.strip()
        if teacher_name[0] in WorkloadType:
            load_type = WorkloadType(teacher_name[0])
            teacher_name = teacher_name[1:].strip()
        else:
            load_type = WorkloadType.MAIN
        cell = ws.cell(row=row, column=lesson_col)
        lesson_name, *subgroup_names = cell.value.split(', п/г ')
        subgroup_name = subgroup_names[-1] if subgroup_names else ''
        cell = ws.cell(row=row, column=group_col)
        group = cell.value
        cell = ws.cell(row=row, column=teachtype_col)
        edu_type = EducationType(cell.value) if cell.value in EducationType else None
        cell = ws.cell(row=row, column=edu_level_col)
        edu_level = cell.value
        cell = ws.cell(row=row, column=activity_col)
        activity = Activity(cell.value) if cell.value in Activity else cell.value
        cell = ws.cell(row=row, column=exam_col)
        exam = ExamType(cell.value) if cell.value in ExamType else None
        cell = ws.cell(row=row, column=year_col)
        syear, _, ssemester = str(cell.value).partition('/')
        cell = ws.cell(row=row, column=students_col)
        student_count = int(cell.value)
        cell = ws.cell(row=row, column=weeks_col)
        week_count = int(cell.value) if cell.value else None
        cell = ws.cell(row=row, column=worktime_col)
        worktime = float(cell.value) if cell.value else 0.0
        cell = ws.cell(row=row, column=othertime_col)
        othertime = float(cell.value) if cell.value else 0.0

        loadunit = LoadUnit(
            course=lesson_name,
            group=group,
            subgroup=int(subgroup_name) if subgroup_name else None,
            education_type=edu_type,
            education_level=edu_level,
            activity=activity,
            exam=exam,
            year=int(syear),
            semester=int(ssemester) if ssemester else None,
            student_count=student_count,
            week_count=week_count,
            worktime_hours=worktime,
            other_hours=othertime
        )
        teacher_workload[teacher_name][load_type].append(loadunit)
        row += 1
    for workload in teacher_workload.values():
        for units in workload.values():
            units.sort(key=lambda unit: unit.course)
    teacher_workload.pop('', None)
    return teacher_workload
