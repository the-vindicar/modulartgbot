"""Работа с шаблонами документов, опираясь на полученные от прасеров данные."""
import typing as t
from collections import defaultdict
import datetime
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from .workload_parsing import TeacherWorkload, LoadUnit, WorkloadType, Activity, EducationType, ExamType
from .timeplan_parsing import GroupPlan, TimePlanActivity


__all__ = ['fill_template']


def split_workload(workload: TeacherWorkload) -> tuple[TeacherWorkload, TeacherWorkload, TeacherWorkload]:
    """Разбивает нагрузку на осенний семестр, весенний семестр и неопределённую.
    Нагрузка, не привязанная к семестру, будет отнесена на весенний семестр.
    У заочников триместры 1 и 2 относятся к осени, а 3 - к весне."""
    autumn: TeacherWorkload = defaultdict(list)
    spring: TeacherWorkload = defaultdict(list)
    unknown: TeacherWorkload = defaultdict(list)
    for loadtype, units in workload.items():
        for unit in units:
            # определяем семестр, к которому относится нагрузка
            if unit.education_type == EducationType.OFFHOURS and unit.semester is not None:
                semester_code = (unit.semester - 1) % 3 + 1
                target = autumn if semester_code in (1, 2) else spring
            elif unit.education_type == EducationType.DAYHOURS and unit.semester is not None:
                semester_code = (unit.semester - 1) % 2 + 1
                target = autumn if semester_code == 1 else spring
            elif unit.education_type == EducationType.DAYHOURS and unit.semester is None:
                target = spring
            else:
                target = unknown
            target[loadtype].append(unit)
    return autumn, spring, unknown


def set_dates(ws: Worksheet, year: int, start_month: int) -> list[datetime.datetime]:
    """Задаёт столбцы месяцев."""
    dates = []
    for i, cell in enumerate(ws['J4:N4'][0]):
        cell: Cell
        month = start_month + i
        dt = datetime.datetime(year=year + int(month > 12), month=(month - 1) % 12 + 1, day=1)
        cell.value = dt
        dates.append(dt)
    return dates


def combine_groups(groups: list[str], hours: list[float], exam_hours: list[float]) -> tuple[str, str, str]:
    """Соединяет список групп с указанными часами в более компактное представление.
    Также объединяет часы в текст формулы Excel. Например, вызов:

    ``combine_groups(['23-ИСбо-1а', '23-ИИбо-4', '23-ИСбо-1б'], [12.0, 13.0, 14.0], [2.0, 1.5, 1.8])``

    Вернёт следующий результат:

    ``('23-ИСбо-1а,1б; 23-ИИбо-4', '=12.0+14.0+13.0', '=2.0+1.8+1.5')``

    Обратите внимание, что порядок групп и порядок часов изменился одинаковым образом.
    """
    data: list[tuple[str, str, float, float]] = list()
    for group, h, eh in zip(groups, hours, exam_hours):
        spec, _, groupidx = group.rpartition('-')
        data.append((spec, groupidx, h, eh))
    data.sort(key=lambda item: item[0:2])
    current_spec = None
    group = ''
    hour_formula = ''
    exam_formula = ''
    for spec, groupidx, h, eh in data:
        if current_spec == spec:
            group += f',{groupidx}'
        else:
            if current_spec is not None:
                group += '; '
            group += f'{spec}-{groupidx}'
            current_spec = spec
        hour_formula += f'+{h}'
        exam_formula += f'+{eh}'
    return group, '='+hour_formula[1:], '='+exam_formula[1:]


def fill_template(template: Path, year: int, workload: TeacherWorkload,
                  plans: t.Collection[GroupPlan] = tuple()) -> Workbook:
    """Заполняет шаблон нагрузки на указанный год, используя указанные сведения.
    :param template: Путь к файлу с заполняемым шаблоном.
    :param year: Год начала курса (меньший из двух).
    :param workload: Нагрузка преподавателя.
    :param plans: Коллекция доступных план-графиков для групп."""
    LOAD_NAMES = {
        WorkloadType.MAIN: 'основная',
        WorkloadType.HOURLY: 'почасовая',
        WorkloadType.INTERNAL: 'совмещение',
    }
    ACTIVITY_NAMES = {
        Activity.LECTURE: 'лекция',
        Activity.PRACTICAL: 'практическая',
        Activity.LAB: 'лабораторная',
        Activity.YEARLYTASK: 'к/р',
        Activity.YEARLYPROJECT: 'к/п',
        Activity.INTERNSHIP: 'практика',
        Activity.GRADPROJECT: 'ВКР',
        Activity.GRADCOMITTEE: 'ГАК',
        Activity.MAGGRADPROJECT: 'магистр.дисс.',
        Activity.MAGISTER: 'магистр',
    }
    EXAM_NAMES = {
        ExamType.EXAM: 'экзамен',
        ExamType.PASS: 'зачёт',
        ExamType.DIFFPASS: 'дифф.зачёт',
    }
    GROUP_NAMES = ['а', 'б']
    ACTIVITY_MAP: dict[Activity, tuple[TimePlanActivity, ...]] = {
        Activity.INTERNSHIP: (TimePlanActivity.INDUSTRY_PRACTICE, TimePlanActivity.STUDY_PRACTICE),
        Activity.GRADPROJECT: (TimePlanActivity.GRAD_PROJECT,),
    }
    wb = openpyxl.load_workbook(template, rich_text=True, data_only=False)
    workload = {lt: units.copy() for lt, units in workload.items()}
    if WorkloadType.G in workload:  # Я не знаю, что такое g, так что обозначаю её как обычную нагрузку
        workload.setdefault(WorkloadType.MAIN, []).extend(workload[WorkloadType.G])
        del workload[WorkloadType.G]
    autumn, spring, unknown = split_workload(workload)
    autumn_s = wb['Осень']
    autumn_dates = set_dates(autumn_s, year, 9)
    spring_s = wb['Весна']
    spring_dates = set_dates(spring_s, year+1, 2)
    pages: list[tuple[TeacherWorkload, Worksheet, list[datetime.datetime], str]] = [
        (autumn, autumn_s, autumn_dates, 'autumn'),
        (spring, spring_s, spring_dates, 'spring'),
    ]
    if unknown and any(unknown.values()):
        unknown_s = wb.copy_worksheet(autumn_s)
        unknown_s.title = 'НЕ ОПРЕДЕЛЕНО'
        pages.append((unknown, unknown_s, autumn_dates, ''))
    for data, sheet, dates, pageid in pages:
        if sheet.freeze_panes:
            cell: Cell = sheet[sheet.freeze_panes]
            row, startcol = cell.row, cell.column
        else:
            row, startcol = 2, 1

        for loadtype in (WorkloadType.MAIN, WorkloadType.INTERNAL, WorkloadType.HOURLY):
            load = data[loadtype]
            if not load:
                continue
            load.sort(key=lambda u: (u.education_level, u.course, u.activity))
            current_key = None
            start: t.Optional[LoadUnit] = None
            groups = []
            hours = []
            exam_hours = []
            plan = None
            for unit in load:
                plan = None
                for p in plans:
                    if p.group_code == unit.group:
                        plan = p
                        break
                if current_key != (unit.education_level, unit.course, unit.activity):
                    if current_key is not None:  # мы не в самом начале цикла?
                        # добавляем элемент нагрузки в шаблон
                        if sum(exam_hours) > 0 and sum(hours) == 0:
                            hours, exam_hours = exam_hours, hours
                        groupnames, hourformula, examformula = combine_groups(groups, hours, exam_hours)
                        if sum(hours) > 0:
                            sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                            sheet.cell(row=row, column=startcol+1, value=start.course)
                            sheet.cell(row=row, column=startcol+2,
                                       value=ACTIVITY_NAMES.get(start.activity, start.activity))
                            sheet.cell(row=row, column=startcol+3, value=groupnames)
                            plan_activity = ACTIVITY_MAP.get(start.activity, tuple())
                            if plan is None or not pageid or not plan_activity:
                                startdate, enddate = dates[0], dates[-1] - datetime.timedelta(days=1)
                            else:
                                startdate, enddate = plan.get_interval(pageid, *plan_activity)
                            sheet.cell(row=row, column=startcol + 4).value = startdate
                            sheet.cell(row=row, column=startcol + 5).value = enddate
                            sheet.cell(row=row, column=startcol+6, value=hourformula)
                            row += 1
                        if sum(exam_hours) > 0 and start.exam is not None:  # есть экзамен или зачёт
                            sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                            sheet.cell(row=row, column=startcol+1, value=start.course)
                            sheet.cell(row=row, column=startcol+2,
                                       value=EXAM_NAMES.get(start.exam, start.exam.value))
                            if plan is None or not pageid:
                                startdate = dates[-1]
                                enddate = dates[-1].replace(month=dates[-1].month + 1) - datetime.timedelta(days=1)
                            else:
                                startdate, enddate = plan.get_interval(pageid, TimePlanActivity.EXAM)
                            sheet.cell(row=row, column=startcol + 4).value = startdate
                            sheet.cell(row=row, column=startcol + 5).value = enddate
                            sheet.cell(row=row, column=startcol+3, value=groupnames)
                            sheet.cell(row=row, column=startcol+6, value=examformula)
                            row += 1
                    # готовимся аккумулировать данные для следующего.
                    current_key = (unit.education_level, unit.course, unit.activity)
                    start = unit
                    groups = []
                    hours = []
                    exam_hours = []
                # аккумулируем данные для очередного элемента
                hours.append(unit.worktime_hours)
                exam_hours.append(unit.other_hours)
                # курсовые и т.п. обычно прописаны на каждого студента
                if unit.student_count > 1 and unit.subgroup is not None and 1 <= unit.subgroup <= len(GROUP_NAMES):
                    groups.append(unit.group + GROUP_NAMES[unit.subgroup-1])
                else:
                    groups.append(unit.group)
            # последний элемент списка ещё не добавлен - добавляем
            if current_key is not None:  # у нас была нагрузка?
                # добавляем элемент нагрузки в шаблон
                if sum(exam_hours) > 0 and sum(hours) == 0:
                    hours, exam_hours = exam_hours, hours
                groupnames, hourformula, examformula = combine_groups(groups, hours, exam_hours)
                if sum(hours) > 0:
                    sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                    sheet.cell(row=row, column=startcol+1, value=start.course)
                    sheet.cell(row=row, column=startcol+2, value=ACTIVITY_NAMES.get(start.activity, start.activity))
                    sheet.cell(row=row, column=startcol+3, value=groupnames)
                    plan_activity = ACTIVITY_MAP.get(start.activity, tuple())
                    if plan is None or not pageid or not plan_activity:
                        startdate, enddate = dates[0], dates[-1] - datetime.timedelta(days=1)
                    else:
                        startdate, enddate = plan.get_interval(pageid, *plan_activity)
                    sheet.cell(row=row, column=startcol + 4).value = startdate
                    sheet.cell(row=row, column=startcol + 5).value = enddate
                    sheet.cell(row=row, column=startcol+6, value=hourformula)
                    row += 1
                if sum(exam_hours) > 0 and start.exam is not None:  # есть экзамен или зачёт
                    sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                    sheet.cell(row=row, column=startcol+1, value=start.course)
                    sheet.cell(row=row, column=startcol+2, value=EXAM_NAMES.get(start.exam, start.exam.value))
                    if plan is None or not pageid:
                        startdate = dates[-1]
                        enddate = dates[-1].replace(month=dates[-1].month + 1) - datetime.timedelta(days=1)
                    else:
                        startdate, enddate = plan.get_interval(pageid, TimePlanActivity.EXAM)
                    sheet.cell(row=row, column=startcol + 4).value = startdate
                    sheet.cell(row=row, column=startcol + 5).value = enddate
                    sheet.cell(row=row, column=startcol+3, value=groupnames)
                    sheet.cell(row=row, column=startcol+6, value=examformula)
                    row += 1
            # пропускаем несколько строк, оставляя их для правок
            row += 3
    return wb
