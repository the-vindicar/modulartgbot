"""Реализует работу с Excel файлами."""
from collections import defaultdict
import dataclasses
import datetime
from enum import StrEnum
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


__all__ = [
    'WorkloadType', 'Activity', 'EducationType', 'ExamType', 'LoadUnit',
    'TeacherWorkload', 'parse_workload', 'fill_template'
]


class WorkloadType(StrEnum):
    """Типы занятости преподавателя (штат, почасовка, совмещение и т.п.)"""
    MAIN = 'ш'
    HOURLY = 'h'
    INTERNAL = 'i'
    G = 'g'
    _order_ = ('MAIN', 'HOURLY', 'INTERNAL', 'G')


class Activity(StrEnum):
    """Виды деятельности"""
    LECTURE = 'Лек'
    PRACTICAL = 'Пр'
    LAB = 'Лаб'
    YEARLYTASK = 'КР'
    YEARLYPROJECT = 'КП'
    INTERNSHIP = 'Прак'
    GRADPROJECT = 'ВКР'
    MAGGRADPROJECT = 'ДРМ'
    MAGISTER = 'РукМаг'
    GRADCOMITTEE = 'ГАК'
    _order_ = ('LECTURE', 'PRACTICAL', 'LAB', 'YEARLYTASK', 'YEARLYPROJECT', 'INTERNSHIP',
               'GRADPROJECT', 'MAGGRADPROJECT', 'MAGISTER', 'GRADCOMITTEE')


class EducationType(StrEnum):
    """Тип обучения: очное, заочное и т.п."""
    DAYHOURS = 'Очная форма'
    OFFHOURS = 'Заочная форма'
    COMBINED = 'Очно-заочная форма'


class ExamType(StrEnum):
    """Тип контроля: зачёт, экзамен и т.п."""
    PASS = 'Зч'
    DIFFPASS = 'ЗчО'
    EXAM = 'Эк'


@dataclasses.dataclass
class LoadUnit:
    """Описание одного вида деятельности преподавателя."""
    course: str
    group: str
    subgroup: int | None
    education_type: EducationType | None
    education_level: str
    activity: Activity | str
    exam: ExamType | None
    year: int
    semester: int | None
    student_count: int
    week_count: int | None
    worktime_hours: float
    other_hours: float


TeacherWorkload = dict[WorkloadType, list[LoadUnit]]


def find_column(ws: Worksheet, *targets: tuple[str, int]) -> int:
    """Находит столбец, содержащий в заданной строке заданное значение, и возвращает его номер.
    Выбрасывает IndexError, если такого значения нет.

    :param ws: Лист, на котором ищем значения.
    :param targets: Последовательность пар "значение-строка". Засчитывается соответствие любой паре."""
    for c in range(1, ws.max_column + 1):
        for value, row in targets:
            cell = ws.cell(row=row, column=c)
            if cell.value and str(cell.value) == value:
                return c
    raise IndexError()


def parse_workload(ws: Worksheet) -> dict[str, TeacherWorkload]:
    """Разделяет нагрузку по преподавателям."""
    lesson_col = find_column(ws, ('6', 5))
    year_col = find_column(ws, ('8', 5))
    group_col = find_column(ws, ('Группа', 1), ('9', 5))
    students_col = find_column(ws, ('Кол-во студентов', 1), ('10', 5))
    weeks_col = find_column(ws, ('Недель', 1), ('11', 5))
    activity_col = find_column(ws, ('Вид занятий', 1), ('12', 5))
    exam_col = find_column(ws, ('Виды контроля', 1), ('14', 5))
    worktime_col = find_column(ws, ('23', 5))
    othertime_col = find_column(ws, ('24', 5))
    teacher_col = find_column(ws, ('Преподаватель', 1), ('27', 5))
    edu_level_col = find_column(ws, ('Уровень', 1), ('45', 5))
    teachtype_col = find_column(ws, ('Форма обучения', 1), ('46', 5))

    teacher_workload = defaultdict(lambda: defaultdict(list))
    header_max_row = max(cellrange.max_row for cellrange in ws.merged_cells.ranges if cellrange.min_row == 1)
    row = header_max_row + 2
    while True:
        if not ws.cell(row=row, column=1).value:
            break
        cell = ws.cell(row=row, column=teacher_col)
        if cell.data_type != 's' or not cell.value:
            row += 1
            continue
        teacher_name = cell.value.strip()
        if teacher_name[0] in WorkloadType:
            load_type = WorkloadType(teacher_name[0])
            teacher_name = teacher_name[1:].strip()
        else:
            load_type = WorkloadType.MAIN
        cell = ws.cell(row=row, column=lesson_col)
        lesson_name, _, subgroup_name = cell.value.partition(', п/г ')
        cell = ws.cell(row=row, column=group_col)
        group = cell.value
        cell = ws.cell(row=row, column=teachtype_col)
        edu_type = EducationType(cell.value) if cell.value in EducationType else None
        cell = ws.cell(row=row, column=edu_level_col)
        edu_level = cell.value
        cell = ws.cell(row=row, column=activity_col)
        activity = Activity(cell.value) if cell.value in Activity else cell.value
        cell = ws.cell(row=row, column=exam_col)
        exam = ExamType(cell.value) if cell.value in ExamType else None
        cell = ws.cell(row=row, column=year_col)
        syear, _, ssemester = str(cell.value).partition('/')
        cell = ws.cell(row=row, column=students_col)
        student_count = int(cell.value)
        cell = ws.cell(row=row, column=weeks_col)
        week_count = int(cell.value) if cell.value else None
        cell = ws.cell(row=row, column=worktime_col)
        worktime = float(cell.value) if cell.value else 0.0
        cell = ws.cell(row=row, column=othertime_col)
        othertime = float(cell.value) if cell.value else 0.0

        loadunit = LoadUnit(
            course=lesson_name,
            group=group,
            subgroup=int(subgroup_name) if subgroup_name else None,
            education_type=edu_type,
            education_level=edu_level,
            activity=activity,
            exam=exam,
            year=int(syear),
            semester=int(ssemester) if ssemester else None,
            student_count=student_count,
            week_count=week_count,
            worktime_hours=worktime,
            other_hours=othertime
        )
        teacher_workload[teacher_name][load_type].append(loadunit)
        row += 1
    for workload in teacher_workload.values():
        for units in workload.values():
            units.sort(key=lambda unit: unit.course)
    teacher_workload.pop('', None)
    return teacher_workload


def split_workload(workload: TeacherWorkload) -> tuple[TeacherWorkload, TeacherWorkload, TeacherWorkload]:
    """Разбивает нагрузку на осенний семестр, весенний семестр и неопределённую.
    Нагрузка, не привязанная к семестру, будет отнесена на весенний семестр.
    У заочников триместры 1 и 2 относятся к осени, а 3 - к весне."""
    autumn: TeacherWorkload = defaultdict(list)
    spring: TeacherWorkload = defaultdict(list)
    unknown: TeacherWorkload = defaultdict(list)
    for loadtype, units in workload.items():
        for unit in units:
            # определяем семестр, к которому относится нагрузка
            if unit.education_type == EducationType.OFFHOURS and unit.semester is not None:
                semester_code = (unit.semester - 1) % 3 + 1
                target = autumn if semester_code in (1, 2) else spring
            elif unit.education_type == EducationType.DAYHOURS and unit.semester is not None:
                semester_code = (unit.semester - 1) % 2 + 1
                target = autumn if semester_code == 1 else spring
            elif unit.education_type == EducationType.DAYHOURS and unit.semester is None:
                target = spring
            else:
                target = unknown
            target[loadtype].append(unit)
    return autumn, spring, unknown


def set_dates(ws: Worksheet, year: int, start_month: int) -> list[datetime.datetime]:
    """Задаёт столбцы месяцев."""
    dates = []
    for i, cell in enumerate(ws['J4:N4'][0]):
        cell: Cell
        month = start_month + i
        dt = datetime.datetime(year=year + int(month > 12), month=(month - 1) % 12 + 1, day=1)
        cell.value = dt
        dates.append(dt)
    return dates


def fix_groups(groups: set[str]) -> str:
    """Преобразует строку с перечислением групп.
    :param groups: Множество строк вида "21-ИСбо-1" или "21-ИИмз-1".
    :returns: Строка вида "21-ИСбо-1,2,4; 21-ИИбо-1".
    """
    splits: list[tuple[str, str, str]] = [g.rpartition('-') for g in groups]
    splits.sort(key=lambda item: item[0])
    result = ''
    last_prefix = None
    for prefix, _, number in splits:
        if prefix != last_prefix:
            if last_prefix is not None:
                result += '; '
            result += f'{prefix}-{number}'
            last_prefix = prefix
        else:
            result += f',{number}'
    return result


def combine_groups(groups: list[str], hours: list[float], exam_hours: list[float]) -> tuple[str, str, str]:
    """Соединяет список групп с указанными часами в более компактное представление.
    Также объединяет часы в текст формулы Excel. Например, вызов:

    ``combine_groups(['23-ИСбо-1а', '23-ИИбо-4', '23-ИСбо-1б'], [12.0, 13.0, 14.0], [2.0, 1.5, 1.8])``

    Вернёт следующий результат:

    ``('23-ИСбо-1а,1б; 23-ИИбо-4', '=12.0+14.0+13.0', '=2.0+1.8+1.5')``

    Обратите внимание, что порядок групп и порядок часов изменился одинаковым образом.
    """
    data: list[tuple[str, str, float, float]] = list()
    for group, h, eh in zip(groups, hours, exam_hours):
        spec, _, groupidx = group.rpartition('-')
        data.append((spec, groupidx, h, eh))
    data.sort(key=lambda item: item[0:2])
    current_spec = None
    group = ''
    hour_formula = ''
    exam_formula = ''
    for spec, groupidx, h, eh in data:
        if current_spec == spec:
            group += f',{groupidx}'
        else:
            if current_spec is not None:
                group += '; '
            group += f'{spec}-{groupidx}'
            current_spec = spec
        hour_formula += f'+{h}'
        exam_formula += f'+{eh}'
    return group, '='+hour_formula[1:], '='+exam_formula[1:]


def fill_template(template: Path, year: int, workload: TeacherWorkload) -> Workbook:
    """Заполняет шаблон нагрузки на указанный год, используя указанные сведения."""
    LOAD_NAMES = {
        WorkloadType.MAIN: 'основная',
        WorkloadType.HOURLY: 'почасовая',
        WorkloadType.INTERNAL: 'совмещение',
    }
    ACTIVITY_NAMES = {
        Activity.LECTURE: 'лекция',
        Activity.PRACTICAL: 'практическая',
        Activity.LAB: 'лабораторная',
        Activity.YEARLYTASK: 'к/р',
        Activity.YEARLYPROJECT: 'к/п',
        Activity.INTERNSHIP: 'практика',
        Activity.GRADPROJECT: 'ВКР',
        Activity.GRADCOMITTEE: 'ГАК',
        Activity.MAGGRADPROJECT: 'магистр.дисс.',
        Activity.MAGISTER: 'магистр',
    }
    EXAM_NAMES = {
        ExamType.EXAM: 'экзамен',
        ExamType.PASS: 'зачёт',
        ExamType.DIFFPASS: 'дифф.зачёт',
    }
    GROUP_NAMES = 'аб'
    wb = openpyxl.load_workbook(template, rich_text=True, data_only=False)
    workload = {lt: units.copy() for lt, units in workload.items()}
    if WorkloadType.G in workload:  # Я не знаю, что такое g, так что обозначаю её как обычную нагрузку
        workload.setdefault(WorkloadType.MAIN, []).extend(workload[WorkloadType.G])
        del workload[WorkloadType.G]
    autumn, spring, unknown = split_workload(workload)
    autumn_s = wb['Осень']
    autumn_dates = set_dates(autumn_s, year, 9)
    spring_s = wb['Весна']
    spring_dates = set_dates(spring_s, year+1, 2)
    pages: list[tuple[TeacherWorkload, Worksheet, list[datetime.datetime]]] = [
        (autumn, autumn_s, autumn_dates),
        (spring, spring_s, spring_dates),
    ]
    if unknown and any(unknown.values()):
        unknown_s = wb.copy_worksheet(autumn_s)
        unknown_s.title = 'НЕ ОПРЕДЕЛЕНО'
        pages.append((unknown, unknown_s, autumn_dates))
    for data, sheet, dates in pages:
        if sheet.freeze_panes:
            cell: Cell = sheet[sheet.freeze_panes]
            row, startcol = cell.row, cell.column
        else:
            row, startcol = 2, 1

        for loadtype in (WorkloadType.MAIN, WorkloadType.INTERNAL, WorkloadType.HOURLY):
            load = data[loadtype]
            if not load:
                continue
            load.sort(key=lambda unit: (unit.education_level, unit.course, unit.activity))
            current_key = None
            start: LoadUnit | None = None
            groups = []
            hours = []
            exam_hours = []
            for unit in load:
                if current_key != (unit.education_level, unit.course, unit.activity):
                    if current_key is not None:  # мы не в самом начале цикла?
                        # добавляем элемент нагрузки в шаблон
                        if sum(exam_hours) > 0 and sum(hours) == 0:
                            hours, exam_hours = exam_hours, hours
                        groupnames, hourformula, examformula = combine_groups(groups, hours, exam_hours)
                        if sum(hours) > 0:
                            sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                            sheet.cell(row=row, column=startcol+1, value=start.course)
                            sheet.cell(row=row, column=startcol+2,
                                       value=ACTIVITY_NAMES.get(start.activity, start.activity))
                            sheet.cell(row=row, column=startcol+3, value=groupnames)
                            sheet.cell(row=row, column=startcol+4).value = dates[0]
                            sheet.cell(row=row, column=startcol+5).value = dates[-1] - datetime.timedelta(days=1)
                            sheet.cell(row=row, column=startcol+6, value=hourformula)
                            row += 1
                        if sum(exam_hours) > 0 and start.exam is not None:  # есть экзамен или зачёт
                            sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                            sheet.cell(row=row, column=startcol+1, value=start.course)
                            sheet.cell(row=row, column=startcol+2,
                                       value=EXAM_NAMES.get(start.exam, start.exam.value))
                            sheet.cell(row=row, column=startcol+4).value = dates[-1]
                            sheet.cell(row=row, column=startcol+5).value = (
                                    dates[-1].replace(month=dates[-1].month + 1) - datetime.timedelta(days=1)
                            )
                            sheet.cell(row=row, column=startcol+3, value=groupnames)
                            sheet.cell(row=row, column=startcol+6, value=examformula)
                            row += 1
                    # готовимся аккумулировать данные для следующего.
                    current_key = (unit.education_level, unit.course, unit.activity)
                    start = unit
                    groups = []
                    hours = []
                    exam_hours = []
                # аккумулируем данные для очередного элемента
                hours.append(unit.worktime_hours)
                exam_hours.append(unit.other_hours)
                # курсовые и т.п. обычно прописаны на каждого студента
                if unit.student_count > 1 and unit.subgroup is not None:
                    groups.append(unit.group + GROUP_NAMES[unit.subgroup-1])
                else:
                    groups.append(unit.group)
            # последний элемент списка ещё не добавлен - добавляем
            if current_key is not None:  # у нас была нагрузка?
                # добавляем элемент нагрузки в шаблон
                if sum(exam_hours) > 0 and sum(hours) == 0:
                    hours, exam_hours = exam_hours, hours
                groupnames, hourformula, examformula = combine_groups(groups, hours, exam_hours)
                if sum(hours) > 0:
                    sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                    sheet.cell(row=row, column=startcol+1, value=start.course)
                    sheet.cell(row=row, column=startcol+2, value=ACTIVITY_NAMES.get(start.activity, start.activity))
                    sheet.cell(row=row, column=startcol+3, value=groupnames)
                    sheet.cell(row=row, column=startcol+4).value = dates[0]
                    sheet.cell(row=row, column=startcol+5).value = dates[-1] - datetime.timedelta(days=1)
                    sheet.cell(row=row, column=startcol+6, value=hourformula)
                    row += 1
                if sum(exam_hours) > 0 and start.exam is not None:  # есть экзамен или зачёт
                    sheet.cell(row=row, column=startcol+0, value=LOAD_NAMES[loadtype])
                    sheet.cell(row=row, column=startcol+1, value=start.course)
                    sheet.cell(row=row, column=startcol+2, value=EXAM_NAMES.get(start.exam, start.exam.value))
                    sheet.cell(row=row, column=startcol+4).value = dates[-1]
                    sheet.cell(row=row, column=startcol+5).value = (
                            dates[-1].replace(month=dates[-1].month+1) - datetime.timedelta(days=1)
                    )
                    sheet.cell(row=row, column=startcol+3, value=groupnames)
                    sheet.cell(row=row, column=startcol+6, value=examformula)
                    row += 1
            # пропускаем несколько строк, оставляя их для правок
            row += 3
    return wb
